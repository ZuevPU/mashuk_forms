import hashlib
import hmac
import json
import logging
import os
import re
import time
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Optional
from urllib.parse import quote

import psycopg
from fastapi import APIRouter, Cookie, HTTPException, Request
from fastapi.responses import JSONResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger("mashuk.admin")
UPLOAD_NAME = re.compile(
    r"^[0-9a-fA-F]{32}_(portfolio|consent)\.[A-Za-z0-9]{1,8}$"
)
MEDIA_TYPES = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".doc": "application/msword",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}
COOKIE = "mashuk_admin"
ROOT = Path(__file__).resolve().parent
SORTABLE = {
    "id": "id",
    "created_at": "created_at",
    "fio_latin": "fio_latin",
    "fio_ru": "fio_ru",
    "email": "email",
    "phone": "phone",
    "country": "country",
    "city": "city",
    "org_name": "org_name",
    "stream": "stream",
}

SORTABLE_INFO = {
    "id": "id",
    "created_at": "created_at",
    "fio_latin": "fio_latin",
    "meal_type": "meal_type",
    "stream": "stream",
    "depart_country": "depart_country",
    "depart_city": "depart_city",
    "visa_needed": "visa_needed",
}

EXCEL_COLUMNS = [
    ("id", "ID"),
    ("created_at", "\u0414\u0430\u0442\u0430 \u043f\u043e\u0434\u0430\u0447\u0438"),
    ("fio_latin", "\u0424\u0418\u041e (passport)"),
    ("fio_ru", "\u0424\u0418\u041e \u0440\u0443\u0441."),
    ("birth_date", "\u0414\u0430\u0442\u0430 \u0440\u043e\u0436\u0434\u0435\u043d\u0438\u044f"),
    ("gender", "\u041f\u043e\u043b"),
    ("country", "\u0421\u0442\u0440\u0430\u043d\u0430"),
    ("city", "\u0413\u043e\u0440\u043e\u0434"),
    ("arrival", "\u041e\u0442\u043a\u0443\u0434\u0430 \u043f\u0440\u0438\u0431\u044b\u0432\u0430\u0435\u0442"),
    ("citizenship", "\u0413\u0440\u0430\u0436\u0434\u0430\u043d\u0441\u0442\u0432\u043e"),
    ("all_citizenships", "\u0412\u0441\u0435 \u0433\u0440\u0430\u0436\u0434\u0430\u043d\u0441\u0442\u0432\u0430"),
    ("phone", "\u0422\u0435\u043b\u0435\u0444\u043e\u043d"),
    ("email", "E-mail"),
    ("messenger", "\u041c\u0435\u0441\u0441\u0435\u043d\u0434\u0436\u0435\u0440"),
    ("messenger_link", "\u0421\u0441\u044b\u043b\u043a\u0430 \u043c\u0435\u0441\u0441\u0435\u043d\u0434\u0436\u0435\u0440\u0430"),
    ("messenger_other", "\u0414\u0440\u0443\u0433\u043e\u0439 \u043c\u0435\u0441\u0441\u0435\u043d\u0434\u0436\u0435\u0440"),
    ("org_name", "\u041e\u0440\u0433\u0430\u043d\u0438\u0437\u0430\u0446\u0438\u044f"),
    ("org_spec", "\u0421\u043f\u0435\u0446\u0438\u0430\u043b\u0438\u0437\u0430\u0446\u0438\u044f"),
    ("org_location", "\u0421\u0442\u0440\u0430\u043d\u0430 \u0438 \u0433\u043e\u0440\u043e\u0434 \u043e\u0440\u0433\u0430\u043d\u0438\u0437\u0430\u0446\u0438\u0438"),
    ("position", "\u0414\u043e\u043b\u0436\u043d\u043e\u0441\u0442\u044c"),
    ("audience", "\u0410\u0443\u0434\u0438\u0442\u043e\u0440\u0438\u044f"),
    ("audience_other", "\u0410\u0443\u0434\u0438\u0442\u043e\u0440\u0438\u044f (\u0434\u0440\u0443\u0433\u043e\u0435)"),
    ("stream", "\u041f\u043e\u0442\u043e\u043a"),
    ("how_learned", "\u041e\u0442\u043a\u0443\u0434\u0430 \u0443\u0437\u043d\u0430\u043b"),
    ("intl_programs", "\u041c\u0435\u0436\u0434\u0443\u043d\u0430\u0440\u043e\u0434\u043d\u044b\u0435 \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u044b"),
    ("intl_details", "\u041c\u0435\u0436\u0434\u0443\u043d\u0430\u0440\u043e\u0434\u043d\u044b\u0435 \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u044b (\u0434\u0435\u0442\u0430\u043b\u0438)"),
    ("ru_programs", "\u0420\u043e\u0441\u0441\u0438\u0439\u0441\u043a\u0438\u0435 \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u044b"),
    ("ru_details", "\u0420\u043e\u0441\u0441\u0438\u0439\u0441\u043a\u0438\u0435 \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u044b (\u0434\u0435\u0442\u0430\u043b\u0438)"),
    ("why", "\u041f\u043e\u0447\u0435\u043c\u0443 \u0432\u0430\u0436\u043d\u043e \u0443\u0447\u0430\u0441\u0442\u0438\u0435"),
    ("coop", "\u041f\u0440\u0435\u0434\u043b\u043e\u0436\u0435\u043d\u0438\u0435 \u043f\u043e \u0441\u043e\u0442\u0440\u0443\u0434\u043d\u0438\u0447\u0435\u0441\u0442\u0432\u0443"),
    ("mentor", "\u041d\u0430\u0441\u0442\u0430\u0432\u043d\u0438\u043a"),
    ("directions", "\u041d\u0430\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u0438\u044f"),
    ("directions_other", "\u041d\u0430\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u0438\u0435 (\u0434\u0440\u0443\u0433\u043e\u0435)"),
    ("address", "\u0410\u0434\u0440\u0435\u0441 \u0440\u0435\u0433\u0438\u0441\u0442\u0440\u0430\u0446\u0438\u0438"),
    ("passport_series", "\u041f\u0430\u0441\u043f\u043e\u0440\u0442: \u0441\u0435\u0440\u0438\u044f"),
    ("passport_number", "\u041f\u0430\u0441\u043f\u043e\u0440\u0442: \u043d\u043e\u043c\u0435\u0440"),
    ("passport_date", "\u041f\u0430\u0441\u043f\u043e\u0440\u0442: \u0434\u0430\u0442\u0430"),
    ("passport_issued", "\u041f\u0430\u0441\u043f\u043e\u0440\u0442: \u043a\u0435\u043c \u0432\u044b\u0434\u0430\u043d"),
    ("portfolio_url", "\u041f\u043e\u0440\u0442\u0444\u043e\u043b\u0438\u043e (\u0441\u0441\u044b\u043b\u043a\u0430)"),
    ("consent_url", "\u0421\u043e\u0433\u043b\u0430\u0441\u0438\u0435 (\u0441\u0441\u044b\u043b\u043a\u0430)"),
]

INFO_EXCEL_COLUMNS = [
    ("id", "ID"),
    ("created_at", "\u0414\u0430\u0442\u0430 \u043f\u043e\u0434\u0430\u0447\u0438"),
    ("fio_latin", "\u0424\u0418\u041e"),
    ("health_limits", "\u041e\u0433\u0440\u0430\u043d\u0438\u0447\u0435\u043d\u0438\u044f \u043f\u043e \u0437\u0434\u043e\u0440\u043e\u0432\u044c\u044e"),
    ("meal_type", "\u0422\u0438\u043f \u043f\u0438\u0442\u0430\u043d\u0438\u044f"),
    ("id_doc_type", "\u0422\u0438\u043f \u0434\u043e\u043a\u0443\u043c\u0435\u043d\u0442\u0430"),
    ("id_doc_series", "\u0421\u0435\u0440\u0438\u044f \u0434\u043e\u043a\u0443\u043c\u0435\u043d\u0442\u0430"),
    ("id_doc_number", "\u041d\u043e\u043c\u0435\u0440 \u0434\u043e\u043a\u0443\u043c\u0435\u043d\u0442\u0430"),
    ("id_doc_issued", "\u0414\u0430\u0442\u0430 \u0432\u044b\u0434\u0430\u0447\u0438"),
    ("id_doc_valid_from", "\u0421\u0440\u043e\u043a \u0434\u0435\u0439\u0441\u0442\u0432\u0438\u044f \u0441"),
    ("id_doc_valid_to", "\u0421\u0440\u043e\u043a \u0434\u0435\u0439\u0441\u0442\u0432\u0438\u044f \u043f\u043e"),
    ("id_doc_issuer", "\u041a\u0435\u043c \u0432\u044b\u0434\u0430\u043d"),
    ("entry_doc_name", "\u0414\u043e\u043a\u0443\u043c\u0435\u043d\u0442 \u0434\u043b\u044f \u0432\u044a\u0435\u0437\u0434\u0430 \u0432 \u0420\u0424"),
    ("entry_doc_series", "\u0412\u044a\u0435\u0437\u0434: \u0441\u0435\u0440\u0438\u044f"),
    ("entry_doc_number", "\u0412\u044a\u0435\u0437\u0434: \u043d\u043e\u043c\u0435\u0440"),
    ("entry_doc_issued", "\u0412\u044a\u0435\u0437\u0434: \u0434\u0430\u0442\u0430 \u0432\u044b\u0434\u0430\u0447\u0438"),
    ("entry_doc_valid_from", "\u0412\u044a\u0435\u0437\u0434: \u0441\u0440\u043e\u043a \u0441"),
    ("entry_doc_valid_to", "\u0412\u044a\u0435\u0437\u0434: \u0441\u0440\u043e\u043a \u043f\u043e"),
    ("entry_doc_issuer", "\u0412\u044a\u0435\u0437\u0434: \u043a\u0435\u043c \u0432\u044b\u0434\u0430\u043d"),
    ("stream", "\u041f\u043e\u0442\u043e\u043a"),
    ("depart_country", "\u0421\u0442\u0440\u0430\u043d\u0430 \u043e\u0442\u044a\u0435\u0437\u0434\u0430"),
    ("depart_city", "\u0413\u043e\u0440\u043e\u0434 \u043e\u0442\u044a\u0435\u0437\u0434\u0430"),
    ("return_ticket", "\u041e\u0431\u0440\u0430\u0442\u043d\u044b\u0439 \u0431\u0438\u043b\u0435\u0442"),
    ("baggage", "\u0411\u0430\u0433\u0430\u0436"),
    ("visa_needed", "\u0412\u0438\u0437\u0430 \u0432 \u0420\u0424"),
    ("transit_visa", "\u0422\u0440\u0430\u043d\u0437\u0438\u0442\u043d\u0430\u044f \u0432\u0438\u0437\u0430"),
    ("agree_tickets", "\u0421\u043e\u0433\u043b\u0430\u0441\u0438\u0435: \u0431\u0438\u043b\u0435\u0442\u044b \u0431\u0435\u0437 \u0432\u043e\u0437\u0432\u0440\u0430\u0442\u0430"),
    ("agree_notice", "\u0421\u043e\u0433\u043b\u0430\u0441\u0438\u0435: \u0443\u0432\u0435\u0434\u043e\u043c\u043b\u0435\u043d\u0438\u0435 \u0437\u0430 10 \u0434\u043d\u0435\u0439"),
    ("agree_truth", "\u0421\u043e\u0433\u043b\u0430\u0441\u0438\u0435: \u0434\u043e\u0441\u0442\u043e\u0432\u0435\u0440\u043d\u043e\u0441\u0442\u044c \u0441\u0432\u0435\u0434\u0435\u043d\u0438\u0439"),
    ("agree_extra_docs", "\u0421\u043e\u0433\u043b\u0430\u0441\u0438\u0435: \u0434\u043e\u043f. \u0434\u043e\u043a\u0443\u043c\u0435\u043d\u0442\u044b"),
    ("agree_refusal", "\u0421\u043e\u0433\u043b\u0430\u0441\u0438\u0435: \u043e\u0442\u043a\u0430\u0437 \u043f\u0440\u0438 \u0442\u0440\u0443\u0434\u043d\u043e\u0441\u0442\u044f\u0445 \u0432\u044a\u0435\u0437\u0434\u0430"),
]


def _password() -> str:
    return os.environ.get("ADMIN_PASSWORD", "").strip()


def _secret() -> bytes:
    raw = os.environ.get("ADMIN_SECRET", "").strip() or _password() or "mashuk-dev"
    return raw.encode("utf-8")


def _sign(payload: str) -> str:
    sig = hmac.new(_secret(), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return payload + "." + sig


def _check_token(token: Optional[str]) -> bool:
    if not token or "." not in token:
        return False
    payload, sig = token.rsplit(".", 1)
    expect = hmac.new(_secret(), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expect, sig):
        return False
    try:
        exp = int(payload)
    except ValueError:
        return False
    return exp >= int(time.time())


def require_admin(mashuk_admin: Optional[str] = Cookie(default=None)):
    if not _password():
        raise HTTPException(503, "ADMIN_PASSWORD is not set")
    if not _check_token(mashuk_admin):
        raise HTTPException(401, "unauthorized")
    return True


def public_origin(request: Request) -> str:
    env = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    if env:
        return env
    proto = (
        request.headers.get("x-forwarded-proto") or request.url.scheme or "https"
    ).split(",")[0].strip()
    if proto != "https" and request.headers.get("x-forwarded-ssl", "").lower() == "on":
        proto = "https"
    host = (
        request.headers.get("x-forwarded-host")
        or request.headers.get("host")
        or request.url.netloc
    ).split(",")[0].strip()
    return proto + "://" + host


def cookie_secure(request: Request) -> bool:
    proto = (
        request.headers.get("x-forwarded-proto") or request.url.scheme or ""
    ).split(",")[0].strip().lower()
    return proto == "https"


def file_token(kind: str, app_id: int) -> str:
    msg = ("file:%s:%s" % (kind, int(app_id))).encode("utf-8")
    return hmac.new(_secret(), msg, hashlib.sha256).hexdigest()


def file_url(request: Request, app_id: int, kind: str) -> str:
    return (
        public_origin(request)
        + "/admin/api/applications/%s/file/%s?t=%s"
        % (int(app_id), kind, file_token(kind, app_id))
    )


def allow_file_access(
    kind: str,
    app_id: int,
    token: Optional[str],
    mashuk_admin: Optional[str],
) -> None:
    expect = file_token(kind, app_id)
    given = str(token or "")
    if given and len(given) == len(expect) and hmac.compare_digest(given, expect):
        return
    require_admin(mashuk_admin)


def primary_upload_dir() -> Path:
    raw = (os.environ.get("UPLOAD_DIR") or "").strip()
    if not raw:
        return (ROOT / "uploads").resolve()
    p = Path(raw)
    return p.resolve() if p.is_absolute() else (ROOT / p).resolve()


def upload_dirs() -> list[Path]:
    raw = (os.environ.get("UPLOAD_DIR") or "").strip()
    candidates = [primary_upload_dir()]
    if raw:
        p = Path(raw)
        candidates.append(p if p.is_absolute() else Path.cwd() / p)
    candidates.extend(
        [
            ROOT / "uploads",
            Path.cwd() / "uploads",
            ROOT / "health" / "uploads",
            Path.cwd().parent / "uploads",
            Path("/app/uploads"),
            Path("/app/health/uploads"),
        ]
    )
    out = []
    seen = set()
    for folder in candidates:
        try:
            resolved = folder.resolve()
        except Exception:
            continue
        key = str(resolved).lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(resolved)
    return out


def find_upload(stored: str) -> Path:
    stored = str(stored or "").strip().replace("\\", "/")
    name = Path(stored).name
    if not name or name in (".", ".."):
        raise HTTPException(404, "file not found")
    folders = upload_dirs()
    candidates = []
    raw = Path(stored)
    if stored and raw.is_absolute():
        candidates.append(raw)
    for folder in folders:
        candidates.append(folder / name)
    seen = set()
    for cand in candidates:
        try:
            resolved = cand.resolve()
        except Exception:
            continue
        key = str(resolved).lower()
        if key in seen:
            continue
        seen.add(key)
        if not resolved.is_file():
            continue
        in_uploads = False
        for folder in folders:
            try:
                resolved.relative_to(folder)
                in_uploads = True
                break
            except ValueError:
                continue
        if in_uploads or UPLOAD_NAME.match(resolved.name):
            return resolved
    log.warning(
        "upload missing name=%s stored=%s dirs=%s",
        name,
        stored,
        [str(d) for d in folders],
    )
    raise HTTPException(404, "file not found")


def send_download(path: Path, name: str) -> Response:
    suffix = path.suffix.lower()
    media = MEDIA_TYPES.get(suffix, "application/octet-stream")
    ascii_name = "file" + suffix
    headers = {
        "Content-Disposition": (
            "attachment; filename=\""
            + ascii_name
            + "\"; filename*=UTF-8''"
            + quote(name)
        ),
        "Cache-Control": "no-store",
        "X-Content-Type-Options": "nosniff",
    }
    return Response(content=path.read_bytes(), media_type=media, headers=headers)


def safe_fio_name(row: dict) -> str:
    raw = str(row.get("fio_latin") or row.get("fio_ru") or "").strip()
    raw = re.sub(r"[^\w\s\-]+", "", raw, flags=re.UNICODE)
    raw = re.sub(r"\s+", "_", raw).strip("._")
    return raw or ("id_%s" % row.get("id"))


def download_name(row: dict, kind: str, suffix: str) -> str:
    base = safe_fio_name(row)
    if kind == "consent":
        return base + (suffix or ".pdf")
    return "%s_portfolio%s" % (base, suffix or ".pdf")


def unique_name(name: str, used: set) -> str:
    stem = Path(name).stem
    suffix = Path(name).suffix
    out = name
    n = 2
    key = out.lower()
    while key in used:
        out = "%s_%s%s" % (stem, n, suffix)
        key = out.lower()
        n += 1
    used.add(key)
    return out


def db():
    url = os.environ.get("DATABASE_URL", "").strip()
    if not url:
        raise HTTPException(500, "DATABASE_URL is not set")
    return psycopg.connect(url)


def as_dicts(cur):
    cols = [d.name for d in cur.description]
    out = []
    for row in cur.fetchall():
        item = {}
        for k, v in zip(cols, row):
            if hasattr(v, "isoformat"):
                item[k] = v.isoformat()
            else:
                item[k] = v
        out.append(item)
    return out


def cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        if isinstance(v, list):
            return ", ".join(str(x) for x in v)
        return json.dumps(v, ensure_ascii=False)
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


def _int_param(raw, default: int, lo: int, hi: int) -> int:
    try:
        n = int(raw)
    except (TypeError, ValueError):
        n = default
    return min(hi, max(lo, n))


def parse_filters(request: Request):
    q = request.query_params
    return {
        "q": (q.get("q") or "").strip(),
        "stream": (q.get("stream") or "").strip(),
        "country": (q.get("country") or "").strip(),
        "gender": (q.get("gender") or "").strip(),
        "date_from": (q.get("date_from") or "").strip(),
        "date_to": (q.get("date_to") or "").strip(),
        "sort": SORTABLE.get(q.get("sort") or "created_at", "created_at"),
        "order": "ASC" if (q.get("order") or "").lower() == "asc" else "DESC",
        "page": _int_param(q.get("page"), 1, 1, 100000),
        "limit": _int_param(q.get("limit"), 50, 10, 100),
    }


def where_sql(f):
    clauses = ["TRUE"]
    args = []
    if f["q"]:
        like = "%" + f["q"] + "%"
        clauses.append(
            "("
            "fio_latin ILIKE %s OR fio_ru ILIKE %s OR email ILIKE %s OR phone ILIKE %s "
            "OR country ILIKE %s OR city ILIKE %s OR org_name ILIKE %s OR citizenship ILIKE %s"
            ")"
        )
        args.extend([like] * 8)
    if f["stream"]:
        clauses.append("stream = %s")
        args.append(f["stream"])
    if f["country"]:
        clauses.append("country ILIKE %s")
        args.append("%" + f["country"] + "%")
    if f["gender"]:
        clauses.append("gender = %s")
        args.append(f["gender"])
    if f["date_from"]:
        clauses.append("created_at::date >= %s")
        args.append(f["date_from"])
    if f["date_to"]:
        clauses.append("created_at::date <= %s")
        args.append(f["date_to"])
    return " AND ".join(clauses), args


def parse_info_filters(request: Request):
    q = request.query_params
    return {
        "q": (q.get("q") or "").strip(),
        "stream": (q.get("stream") or "").strip(),
        "country": (q.get("country") or "").strip(),
        "date_from": (q.get("date_from") or "").strip(),
        "date_to": (q.get("date_to") or "").strip(),
        "sort": SORTABLE_INFO.get(q.get("sort") or "created_at", "created_at"),
        "order": "ASC" if (q.get("order") or "").lower() == "asc" else "DESC",
        "page": _int_param(q.get("page"), 1, 1, 100000),
        "limit": _int_param(q.get("limit"), 50, 10, 100),
    }


def where_info_sql(f):
    clauses = ["TRUE"]
    args = []
    if f["q"]:
        like = "%" + f["q"] + "%"
        clauses.append(
            "("
            "fio_latin ILIKE %s OR depart_country ILIKE %s OR depart_city ILIKE %s "
            "OR stream ILIKE %s OR meal_type ILIKE %s OR id_doc_number ILIKE %s"
            ")"
        )
        args.extend([like] * 6)
    if f["stream"]:
        clauses.append("stream = %s")
        args.append(f["stream"])
    if f["country"]:
        clauses.append("depart_country ILIKE %s")
        args.append("%" + f["country"] + "%")
    if f["date_from"]:
        clauses.append("created_at::date >= %s")
        args.append(f["date_from"])
    if f["date_to"]:
        clauses.append("created_at::date <= %s")
        args.append(f["date_to"])
    return " AND ".join(clauses), args


router = APIRouter()


@router.post("/api/login")
async def login(request: Request):
    pwd = _password()
    if not pwd:
        raise HTTPException(503, "ADMIN_PASSWORD is not set")
    try:
        body = await request.json()
    except Exception:
        body = {}
    given = str((body or {}).get("password") or "")
    given_h = hashlib.sha256(given.encode("utf-8")).digest()
    pwd_h = hashlib.sha256(pwd.encode("utf-8")).digest()
    if not hmac.compare_digest(given_h, pwd_h):
        raise HTTPException(401, "bad password")
    token = _sign(str(int(time.time()) + 7 * 24 * 3600))
    response = JSONResponse({"ok": True})
    response.set_cookie(
        COOKIE,
        token,
        httponly=True,
        samesite="lax",
        max_age=7 * 24 * 3600,
        path="/",
        secure=cookie_secure(request),
    )
    return response


@router.post("/api/logout")
def logout():
    response = JSONResponse({"ok": True})
    response.delete_cookie(COOKIE, path="/")
    return response


@router.get("/api/me")
def me(mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    return {"ok": True}


@router.get("/api/meta")
def meta(request: Request, mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    form = (request.query_params.get("form") or "apply").strip().lower()
    table = "participant_details" if form == "info" else "seminar_applications"
    country_col = "depart_country" if form == "info" else "country"
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            total = cur.fetchone()[0]
            cur.execute(
                f"SELECT DISTINCT stream FROM {table} "
                "WHERE stream IS NOT NULL AND stream <> '' ORDER BY 1"
            )
            streams = [r[0] for r in cur.fetchall()]
            cur.execute(
                f"SELECT DISTINCT {country_col} FROM {table} "
                f"WHERE {country_col} IS NOT NULL AND {country_col} <> '' ORDER BY 1"
            )
            countries = [r[0] for r in cur.fetchall()]
            genders = []
            if form != "info":
                cur.execute(
                    "SELECT DISTINCT gender FROM seminar_applications "
                    "WHERE gender IS NOT NULL AND gender <> '' ORDER BY 1"
                )
                genders = [r[0] for r in cur.fetchall()]
    return {
        "form": form,
        "total": total,
        "streams": streams,
        "countries": countries,
        "genders": genders,
    }


@router.get("/api/applications")
def list_applications(request: Request, mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    f = parse_filters(request)
    where, args = where_sql(f)
    offset = (f["page"] - 1) * f["limit"]
    sql_count = f"SELECT COUNT(*) FROM seminar_applications WHERE {where}"
    sql = (
        "SELECT id, created_at, fio_latin, fio_ru, email, phone, country, city, "
        "org_name, position, stream, gender, citizenship, "
        "(portfolio_path IS NOT NULL AND portfolio_path <> '') AS has_portfolio, "
        "(consent_path IS NOT NULL AND consent_path <> '') AS has_consent "
        f"FROM seminar_applications WHERE {where} "
        f"ORDER BY {f['sort']} {f['order']} NULLS LAST "
        "LIMIT %s OFFSET %s"
    )
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(sql_count, args)
            total = cur.fetchone()[0]
            cur.execute(sql, args + [f["limit"], offset])
            items = as_dicts(cur)
    return {
        "items": items,
        "total": total,
        "page": f["page"],
        "limit": f["limit"],
        "pages": max(1, (total + f["limit"] - 1) // f["limit"]),
        "sort": f["sort"],
        "order": f["order"].lower(),
    }


@router.get("/api/applications/export.xlsx")
def export_xlsx(request: Request, mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    f = parse_filters(request)
    where, args = where_sql(f)
    sql = (
        "SELECT * FROM seminar_applications WHERE {where} "
        "ORDER BY {sort} {order} NULLS LAST"
    ).format(where=where, sort=f["sort"], order=f["order"])
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, args)
            rows = as_dicts(cur)
    for row in rows:
        app_id = int(row["id"])
        row["portfolio_url"] = (
            file_url(request, app_id, "portfolio") if row.get("portfolio_path") else ""
        )
        row["consent_url"] = (
            file_url(request, app_id, "consent") if row.get("consent_path") else ""
        )
    return _xlsx(
        rows,
        EXCEL_COLUMNS,
        "Zayavki",
        "mashuk_zayavki.xlsx",
        link_keys=("portfolio_url", "consent_url"),
    )


@router.get("/api/consents.zip")
def export_consents_zip(request: Request, mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    f = parse_filters(request)
    where, args = where_sql(f)
    sql = (
        "SELECT id, fio_latin, fio_ru, consent_path FROM seminar_applications "
        "WHERE {where} AND consent_path IS NOT NULL AND consent_path <> '' "
        "ORDER BY {sort} {order} NULLS LAST"
    ).format(where=where, sort=f["sort"], order=f["order"])
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, args)
            rows = as_dicts(cur)
    buf = BytesIO()
    used = set()
    added = 0
    missing = []
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            stored = row.get("consent_path") or ""
            try:
                path = find_upload(stored)
            except HTTPException:
                missing.append("%s\t%s" % (row.get("id"), Path(str(stored)).name))
                continue
            name = unique_name(
                download_name(row, "consent", path.suffix.lower() or ".pdf"),
                used,
            )
            zf.write(path, name)
            added += 1
        if missing:
            zf.writestr("missing.txt", "id\tfile\n" + "\n".join(missing) + "\n")
        if added == 0:
            zf.writestr(
                "README.txt",
                (
                    "\u0424\u0430\u0439\u043b\u044b \u0441\u043e\u0433\u043b\u0430\u0441\u0438\u0439 "
                    "\u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u044b \u043d\u0430 \u0434\u0438\u0441\u043a\u0435 "
                    "\u0441\u0435\u0440\u0432\u0435\u0440\u0430.\n"
                    "\u041e\u043d\u0438 \u043c\u043e\u0433\u043b\u0438 \u043f\u0440\u043e\u043f\u0430\u0441\u0442\u044c "
                    "\u043f\u043e\u0441\u043b\u0435 \u0434\u0435\u043f\u043b\u043e\u044f \u0431\u0435\u0437 "
                    "\u043f\u043e\u0441\u0442\u043e\u044f\u043d\u043d\u043e\u0433\u043e \u0445\u0440\u0430\u043d\u0438\u043b\u0438\u0449\u0430.\n"
                ),
            )
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": 'attachment; filename="mashuk_soglasia.zip"',
            "Cache-Control": "no-store",
        },
    )


@router.get("/api/applications/{app_id}")
def get_application(
    app_id: int,
    request: Request,
    mashuk_admin: Optional[str] = Cookie(default=None),
):
    require_admin(mashuk_admin)
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM seminar_applications WHERE id = %s", (app_id,)
            )
            items = as_dicts(cur)
    if not items:
        raise HTTPException(404, "not found")
    item = items[0]
    item["has_portfolio"] = bool(item.get("portfolio_path"))
    item["has_consent"] = bool(item.get("consent_path"))
    item["portfolio_url"] = (
        file_url(request, app_id, "portfolio") if item["has_portfolio"] else ""
    )
    item["consent_url"] = (
        file_url(request, app_id, "consent") if item["has_consent"] else ""
    )
    item.pop("portfolio_path", None)
    item.pop("consent_path", None)
    return item


@router.get("/api/applications/{app_id}/file/{kind}")
def get_file(
    app_id: int,
    kind: str,
    t: Optional[str] = None,
    mashuk_admin: Optional[str] = Cookie(default=None),
):
    if kind not in ("portfolio", "consent"):
        raise HTTPException(404, "unknown file")
    allow_file_access(kind, app_id, t, mashuk_admin)
    col = "portfolio_path" if kind == "portfolio" else "consent_path"
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {col}, fio_latin, fio_ru FROM seminar_applications WHERE id = %s",
                (app_id,),
            )
            row = cur.fetchone()
    if not row or not row[0]:
        raise HTTPException(404, "file not found")
    path = find_upload(row[0])
    name = download_name(
        {"id": app_id, "fio_latin": row[1], "fio_ru": row[2]},
        kind,
        path.suffix.lower() or ".pdf",
    )
    return send_download(path, name)


def _xlsx(rows, columns, sheet_name, filename, link_keys=()):
    links = set(link_keys or ())
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="223F9A")
    link_font = Font(color="0563C1", underline="single")
    for i, (_key, title) in enumerate(columns, 1):
        cell_ref = ws.cell(1, i, title)
        cell_ref.font = header_font
        cell_ref.fill = header_fill
        cell_ref.alignment = Alignment(wrap_text=True, vertical="center")
    for r_i, row in enumerate(rows, 2):
        for c_i, (key, _title) in enumerate(columns, 1):
            value = cell(row.get(key))
            ref = ws.cell(r_i, c_i, value)
            if key in links and value:
                ref.hyperlink = value
                ref.font = link_font
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows)+1)}"
    ws.freeze_panes = "A2"
    for i, (key, _title) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 48 if key in links else 22
    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


@router.get("/api/participants")
def list_participants(request: Request, mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    f = parse_info_filters(request)
    where, args = where_info_sql(f)
    offset = (f["page"] - 1) * f["limit"]
    sql_count = f"SELECT COUNT(*) FROM participant_details WHERE {where}"
    sql = (
        "SELECT id, created_at, fio_latin, meal_type, stream, depart_country, "
        "depart_city, visa_needed, transit_visa, return_ticket, baggage "
        f"FROM participant_details WHERE {where} "
        f"ORDER BY {f['sort']} {f['order']} NULLS LAST "
        "LIMIT %s OFFSET %s"
    )
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(sql_count, args)
            total = cur.fetchone()[0]
            cur.execute(sql, args + [f["limit"], offset])
            items = as_dicts(cur)
    return {
        "items": items,
        "total": total,
        "page": f["page"],
        "limit": f["limit"],
        "pages": max(1, (total + f["limit"] - 1) // f["limit"]),
        "sort": f["sort"],
        "order": f["order"].lower(),
    }


@router.get("/api/participants/export.xlsx")
def export_participants_xlsx(request: Request, mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    f = parse_info_filters(request)
    where, args = where_info_sql(f)
    sql = (
        f"SELECT * FROM participant_details WHERE {where} "
        f"ORDER BY {f['sort']} {f['order']} NULLS LAST"
    )
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, args)
            rows = as_dicts(cur)
    return _xlsx(rows, INFO_EXCEL_COLUMNS, "Uchastniki", "mashuk_uchastniki.xlsx")


@router.get("/api/participants/{item_id}")
def get_participant(item_id: int, mashuk_admin: Optional[str] = Cookie(default=None)):
    require_admin(mashuk_admin)
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM participant_details WHERE id = %s", (item_id,)
            )
            items = as_dicts(cur)
    if not items:
        raise HTTPException(404, "not found")
    item = items[0]
    item.pop("payload_raw", None)
    return item

